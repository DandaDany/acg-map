#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import user-maintained ACG events from 全台ACG活動.xlsx.

The Location column is expected to be written as:
    場地名稱（完整地址）

The importer keeps the venue name outside the final address parentheses and
stores the address inside those parentheses as ``addr`` in manual_extra.json.
"""
import datetime
from paths import path as P
import hashlib
import json
import os
import re
import ast

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = P("全台ACG活動.xlsx")
OUT = P("manual_extra.json")
TOWN_CENTROIDS = P("town_centroids.json")
VENUES = P("venues.json")
CACO_STORES = P("caco_stores.json")
AUTO_EXTRA_FILES = [
    P("venue_extra.json"),
    P("soka_extra.json"),
    P("caco_extra.json"),
    P("cayenne_extra.json"),
]
AUTO_COLLECTORS = [
    P("collect_venues.py"),
    P("collect_public.py"),
]

TODAY = datetime.date.today().strftime("%Y/%m/%d")
CITY = r"(台北市|臺北市|新北市|桃園市|台中市|臺中市|台南市|臺南市|高雄市|基隆市|新竹市|新竹縣|苗栗縣|彰化縣|南投縣|雲林縣|嘉義市|嘉義縣|屏東縣|宜蘭縣|花蓮縣|台東縣|臺東縣|澎湖縣|金門縣|連江縣)"
TOWN_RE = re.compile(CITY + r"([一-龥]{1,3}[區鄉鎮市])")
CITY_RE = re.compile(CITY)
CITY_ALIAS = {
    "台北": "台北市", "臺北": "台北市",
    "台中": "台中市", "臺中": "台中市",
    "台南": "台南市", "臺南": "台南市",
    "高雄": "高雄市", "基隆": "基隆市",
    "新竹": "新竹市", "桃園": "桃園市",
    "屏東": "屏東縣", "宜蘭": "宜蘭縣",
    "花蓮": "花蓮縣", "台東": "台東縣", "臺東": "台東縣",
    "澎湖": "澎湖縣", "金門": "金門縣",
}


def norm(text):
    return (text or "").replace("臺", "台").strip()


def clean_addr(addr):
    return re.sub(r"^\s*\d{3,6}\s*", "", norm(addr))


def compact(text):
    return re.sub(r"[^0-9A-Za-z一-龥]+", "", norm(text).lower())


def loose_compact(text):
    text = compact(text)
    for city in ("台北市", "新北市", "桃園市", "台中市", "台南市", "高雄市", "基隆市", "新竹市", "嘉義市"):
        text = text.replace(compact(city), compact(city[:-1]))
    return text


def fmt_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        value = value.date()
    if isinstance(value, datetime.date):
        return value.strftime("%Y/%m/%d")
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace("-", "/").replace(".", "/")
    m = re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", text)
    if m:
        return f"{int(m.group(1)):04d}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"
    return text


def cell_link(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target.strip()
    value = str(cell.value or "").strip()
    return value if value.startswith(("http://", "https://")) else ""


def parse_location(value):
    raw = norm(str(value or ""))
    if not raw:
        return "", ""
    raw = re.sub(r"[（(]\s*[）)]", "", raw).strip()
    # Use the last parenthetical segment that contains a city/county as address.
    matches = list(re.finditer(r"[（(]([^（）()]+)[）)]", raw))
    for m in reversed(matches):
        inside = clean_addr(m.group(1))
        if CITY_RE.search(inside):
            venue = (raw[:m.start()] + raw[m.end():]).strip()
            venue = re.sub(r"\s+", " ", venue).strip(" -–—")
            return venue or raw, inside
    return raw, ""


def looks_like_addr(text):
    text = clean_addr(str(text or ""))
    return bool(CITY_RE.search(text) and re.search(r"[區鄉鎮市].*[路街道段巷弄號]|[路街道段巷弄號]", text))


def city_of(addr):
    m = CITY_RE.search(addr or "")
    return norm(m.group(1)) if m else ""


def infer_city(*texts):
    joined = " ".join(norm(t) for t in texts if t)
    city = city_of(joined)
    if city:
        return city
    for k, v in CITY_ALIAS.items():
        if k in joined:
            return v
    return ""


def jitter(name, amp=0.012):
    h = int(hashlib.md5(name.encode("utf-8")).hexdigest(), 16)
    return ((h % 1000) / 1000 - 0.5) * 2 * amp, (((h // 1000) % 1000) / 1000 - 0.5) * 2 * amp


def load_centroids():
    raw = json.load(open(TOWN_CENTROIDS, encoding="utf-8"))
    cent = {}
    city_points = {}
    for key, value in raw.items():
        if "|" not in key:
            continue
        city, town = key.split("|", 1)
        city = norm(city)
        town = norm(town)
        cent[(city, town)] = value
        city_points.setdefault(city, []).append(value)
    citycent = {}
    for city, points in city_points.items():
        citycent[city] = [sum(p[0] for p in points) / len(points), sum(p[1] for p in points) / len(points)]
    return cent, citycent


def known_exact_venues():
    out = {}
    if not os.path.exists(VENUES):
        return out
    try:
        venues = json.load(open(VENUES, encoding="utf-8")).get("venues", [])
    except Exception:
        return out
    for v in venues:
        if v.get("loc") == "exact" and v.get("name") and v.get("la") and v.get("lo"):
            out[norm(v["name"])] = {
                "lat": float(v["la"]),
                "lng": float(v["lo"]),
                "city": norm(v.get("city", "")),
                "loc": "exact",
                "addr": clean_addr(v.get("addr", "")),
            }
    return out


def _literal_assigned_value(path, var_name):
    try:
        tree = ast.parse(open(path, encoding="utf-8").read())
    except Exception:
        return None
    for node in tree.body:
        if not isinstance(node, ast.Assign):
            continue
        if not any(isinstance(t, ast.Name) and t.id == var_name for t in node.targets):
            continue
        try:
            return ast.literal_eval(node.value)
        except Exception:
            return None
    return None


def load_auto_managed_venues():
    """場館已由官方/專屬流程維護時，Excel 不再重複匯入。

    來源包含 collect_venues.py / collect_public.py 的設定，以及目前已產生的
    venue_extra/soka_extra/caco_extra/cayenne_extra。這些資料會在 update_all.py 中自動重建。
    """
    names = set()

    for path in AUTO_COLLECTORS:
        venues = _literal_assigned_value(path, "VENUES")
        if isinstance(venues, list):
            for item in venues:
                if isinstance(item, dict) and item.get("key"):
                    names.add(norm(item["key"]))
        elif isinstance(venues, dict):
            names.update(norm(k) for k in venues if k)

    for path in AUTO_EXTRA_FILES:
        if not os.path.exists(path):
            continue
        try:
            data = json.load(open(path, encoding="utf-8"))
        except Exception:
            continue
        if isinstance(data, dict) and isinstance(data.get("venues"), dict):
            names.update(norm(k) for k in data["venues"] if k)
        elif isinstance(data, dict):
            names.update(norm(k) for k in data if k and not str(k).startswith("_"))

    return {n for n in names if n}


def auto_managed_reason(venue, auto_names):
    c = loose_compact(venue)
    if not c:
        return ""
    for name in sorted(auto_names, key=len, reverse=True):
        n = loose_compact(name)
        if len(n) < 4:
            continue
        if c == n or c.startswith(n) or n in c:
            return name
    return ""


def load_caco_cafes():
    if not os.path.exists(CACO_STORES):
        return []
    try:
        data = json.load(open(CACO_STORES, encoding="utf-8"))
    except Exception:
        return []
    cafes = []
    for item in data.get("cafes", []):
        title = norm(item.get("title", ""))
        addr = clean_addr(item.get("addr", ""))
        if title and addr:
            cafes.append({
                "title": title,
                "addr": addr,
                "source": item.get("source", ""),
                "key": compact(title),
            })
    return cafes


def caco_cafe_matches(entry):
    hay = compact(" ".join([entry.get("venue", ""), entry.get("title", "")]))
    return "cacocafe" in hay or "caco咖啡" in hay


def fill_caco_cafe_addr(entry, cafes):
    if entry.get("addr") or not cafes:
        return []
    hay = compact(" ".join([entry.get("venue", ""), entry.get("title", "")]))
    if not caco_cafe_matches(entry):
        return []
    for cafe in sorted(cafes, key=lambda c: len(c["key"]), reverse=True):
        key = cafe["key"]
        if len(key) < 11:
            continue
        if key in hay:
            clone = dict(entry)
            clone["venue"] = cafe["title"]
            clone["addr"] = cafe["addr"]
            if cafe.get("source") and not clone.get("link"):
                clone["link"] = cafe["source"]
            return [clone]

    # If the spreadsheet only says CACO CAFE, treat it as all official CAFE branches.
    out = []
    for cafe in cafes:
        clone = dict(entry)
        clone["venue"] = cafe["title"]
        clone["addr"] = cafe["addr"]
        if cafe.get("source") and not clone.get("link"):
            clone["link"] = cafe["source"]
        out.append(clone)
    return out


def locate(venue, addr, cent, citycent, known):
    key = norm(venue)
    if key in known:
        rec = known[key]
        return rec["lat"], rec["lng"], "exact", rec.get("city") or city_of(addr)
    m = TOWN_RE.search(addr or "")
    if m:
        city, town = norm(m.group(1)), norm(m.group(2))
        if (city, town) in cent:
            la, lo = cent[(city, town)]
            dx, dy = jitter(venue)
            return round(la + dy, 5), round(lo + dx, 5), "區級", city
    city = city_of(addr)
    if city and city in citycent:
        la, lo = citycent[city]
        dx, dy = jitter(venue, 0.03)
        return round(la + dy, 5), round(lo + dx, 5), "市級", city
    city = infer_city(venue, addr)
    if city and city in citycent:
        la, lo = citycent[city]
        dx, dy = jitter(venue, 0.03)
        return round(la + dy, 5), round(lo + dx, 5), "市級", city
    return None


def main():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    required = ["開始日期 / Start Date", "結束日期 / End Date", "活動名稱 / Activity Name", "地點 / Location", "活動連結 / Activity link"]
    missing = [h for h in required if h not in idx]
    if missing:
        raise SystemExit(f"Excel 缺少欄位: {missing}")

    cent, citycent = load_centroids()
    known = known_exact_venues()
    auto_managed = load_auto_managed_venues()
    caco_cafes = load_caco_cafes()
    cat2_col = idx.get("活動類別 / Activity Category")  # 軸B形式欄，可選
    kv_col = idx.get("KV")  # 主視覺欄（可填圖片網址）；有填就用，覆寫自動抓的 KV
    entries = []
    parsed_addr = no_addr = caco_addr = imported_events = skipped_past = continuation_addr = skipped_no_location = skipped_auto_managed = 0

    for row in range(2, ws.max_row + 1):
        title = str(ws.cell(row, idx["活動名稱 / Activity Name"]).value or "").strip()
        loc_text = ws.cell(row, idx["地點 / Location"]).value
        if not title and loc_text and entries and looks_like_addr(loc_text):
            addr = clean_addr(loc_text)
            if not entries[-1].get("addr"):
                entries[-1]["addr"] = addr
                continuation_addr += 1
            continue
        if not title or not loc_text:
            continue
        start = fmt_date(ws.cell(row, idx["開始日期 / Start Date"]).value)
        end = fmt_date(ws.cell(row, idx["結束日期 / End Date"]).value)
        if end and end < TODAY:
            skipped_past += 1
            continue
        venue, addr = parse_location(loc_text)
        if addr:
            parsed_addr += 1
        else:
            no_addr += 1
        link = cell_link(ws.cell(row, idx["活動連結 / Activity link"]))
        raw_cat2 = str(ws.cell(row, cat2_col).value or "").strip() if cat2_col else ""
        img = cell_link(ws.cell(row, kv_col)) if kv_col else ""  # KV 欄填的圖片網址（http/https）
        entries.append({
            "venue": venue,
            "addr": addr,
            "start": start,
            "end": end,
            "title": title,
            "link": link,
            "cat2": raw_cat2,
            "img": img,
        })

    expanded = []
    for entry in entries:
        caco_entries = fill_caco_cafe_addr(entry, caco_cafes)
        if caco_entries:
            expanded.extend(caco_entries)
            caco_addr += len(caco_entries)
        else:
            expanded.append(entry)
    entries = expanded

    # 供 refresh_venues 的「Excel × 文創 重疊 → 改判 ACG」規則使用：匯出所有使用者 Excel 活動標題。
    # 含官方/專屬流程已維護的場館（例如六大文創園區）——這些場館的活動不會另建 manual 場館（避免
    # 重複圖釘、下方 auto_managed 會略過），但標題仍需讓 refresh 端比對重疊，把對應官網活動改判為 ACG。
    overlap_titles = sorted({e["title"] for e in entries if e.get("title")})
    json.dump(overlap_titles, open(P("excel_overlap_titles.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    manual = {}
    for entry in entries:
        venue = entry["venue"]
        managed = auto_managed_reason(venue, auto_managed)
        if managed:
            skipped_auto_managed += 1
            continue
        addr = entry["addr"]
        located = locate(venue, addr, cent, citycent, known)
        if not located:
            skipped_no_location += 1
            continue
        lat, lng, loc, city = located
        link = entry["link"]
        rec = manual.setdefault(venue, {
            "city": city,
            "lat": lat,
            "lng": lng,
            "loc": loc,
            "addr": addr,
            "url": link,
            "ex": [],
        })
        if addr and (not rec.get("addr") or len(addr) > len(rec.get("addr", ""))):
            rec["addr"] = addr
        if link and not rec.get("url"):
            rec["url"] = link
        seen = {e["t"] for e in rec["ex"]}
        title = entry["title"]
        if title not in seen:
            ev = {"t": title, "s": entry["start"], "e": entry["end"], "l": link, "img": entry.get("img", "")}
            if entry.get("cat2"):
                ev["cat2"] = entry["cat2"]
            rec["ex"].append(ev)
            imported_events += 1

    json.dump(dict(sorted(manual.items())), open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(json.dumps({
        "venues": len(manual),
        "events": imported_events,
        "parsed_addr_rows": parsed_addr,
        "caco_cafe_addr_rows": caco_addr,
        "continuation_addr_rows": continuation_addr,
        "no_addr_rows": no_addr,
        "skipped_past": skipped_past,
        "skipped_auto_managed": skipped_auto_managed,
        "skipped_no_location": skipped_no_location,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
