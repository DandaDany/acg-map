#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把 data/manual/全台ACG活動.xlsx 匯出成 diff 友善的純文字鏡像 data/manual/acg_events.json。

動機：xlsx 是二進位檔，GitHub PR diff 看不出改了哪幾列；本檔產生的 JSON 是
「Excel 的忠實文字鏡像」，供 import_acg_excel.py 優先讀取（不存在時仍退回讀 xlsx）。

格式（頂層為陣列，一元素 = 一資料列，依 Excel 第 2 列起）：
  [
    {"KV": null, "類型 / Category": "展覽", ..., "_links": {"活動連結 / Activity link": "https://..."}},
    ...
  ]
規則：
  - 鍵為表頭欄名（第 1 列、str(value).strip()，與 import_acg_excel.py 讀表頭方式一致），
    依欄位順序保留所有欄。
  - 空儲存格 → null；字串/數字原樣保留；datetime → "YYYY-MM-DD HH:MM:SS"、date → "YYYY-MM-DD"
    （import 端 fmt_date 對此字串與原 datetime 產生完全相同的結果）。
  - 儲存格若帶超連結（顯示文字可能 ≠ 連結目標，實測有 5 筆），目標存於該列的
    "_links": {欄名: 目標網址}；import 端 cell_link 會優先取用，等同 openpyxl 的 cell.hyperlink。
  - 可重複執行（冪等）：同一份 xlsx 匯出結果逐位元組相同。

用法：python3 backend/export_excel_to_json.py
"""
import datetime
import json

from openpyxl import load_workbook

from paths import path as P

XLSX = P("全台ACG活動.xlsx")
OUT = P("acg_events.json")


def _json_value(value):
    """把 openpyxl 儲存格值轉成 JSON 可序列化、且 import 端解析結果不變的形式。"""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def export_rows():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    if any(not h for h in headers):
        raise SystemExit(f"表頭含空白欄名，無法以 {{欄名: 值}} 匯出: {headers}")
    if len(set(headers)) != len(headers):
        raise SystemExit(f"表頭欄名重複，無法以 {{欄名: 值}} 匯出: {headers}")

    rows = []
    for r in range(2, ws.max_row + 1):
        obj = {}
        links = {}
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(r, i)
            obj[header] = _json_value(cell.value)
            if cell.hyperlink and cell.hyperlink.target:
                links[header] = cell.hyperlink.target
        if links:
            obj["_links"] = links
        rows.append(obj)
    return headers, rows


def main():
    headers, rows = export_rows()
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=1)
        f.write("\n")
    n_links = sum(1 for r in rows if r.get("_links"))
    print(json.dumps({
        "out": OUT,
        "columns": len(headers),
        "rows": len(rows),
        "rows_with_links": n_links,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
