#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Normalize 全台ACG活動.xlsx location rows.

Some records are entered as two rows:

    row N:   活動名稱 ... 地點
    row N+1:              地址

This script folds the address row into the previous location cell so it becomes:

    地點（地址）

It creates a timestamped backup before saving the workbook in place.
"""
import datetime
from paths import path as P
import json
import os
import re
import shutil

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = P("全台ACG活動.xlsx")
BACKUP_DIR = P("_excel_backups")

CITY = r"(台北市|臺北市|新北市|桃園市|台中市|臺中市|台南市|臺南市|高雄市|基隆市|新竹市|新竹縣|苗栗縣|彰化縣|南投縣|雲林縣|嘉義市|嘉義縣|屏東縣|宜蘭縣|花蓮縣|台東縣|臺東縣|澎湖縣|金門縣|連江縣)"
CITY_RE = re.compile(CITY)


def norm(text):
    return (text or "").replace("臺", "台").strip()


def clean_addr(addr):
    return re.sub(r"^\s*\d{3,6}\s*", "", norm(str(addr or "")))


def looks_like_addr(text):
    text = clean_addr(text)
    return bool(CITY_RE.search(text) and re.search(r"[區鄉鎮市].*[路街道段巷弄號]|[路街道段巷弄號]", text))


def has_address_parens(text):
    text = norm(text)
    for m in re.finditer(r"[（(]([^（）()]+)[）)]", text):
        if CITY_RE.search(clean_addr(m.group(1))):
            return True
    return False


def merge_location(location, addr):
    location = norm(str(location or ""))
    addr = clean_addr(addr)
    if not location or not addr:
        return location
    if has_address_parens(location):
        return location
    # Fill empty parentheses such as "活動會館（）".
    if re.search(r"[（(]\s*[）)]\s*$", location):
        return re.sub(r"[（(]\s*[）)]\s*$", f"（{addr}）", location)
    return f"{location}（{addr}）"


def row_has_meaningful_data(ws, row, ignore_col):
    for col in range(1, ws.max_column + 1):
        if col == ignore_col:
            continue
        value = ws.cell(row, col).value
        if value not in (None, ""):
            return True
    return False


def main():
    wb = load_workbook(XLSX)
    total_merged = 0
    total_cleared = 0
    details = []

    for ws in wb.worksheets:
        headers = [str(c.value or "").strip() for c in ws[1]]
        idx = {h: i + 1 for i, h in enumerate(headers)}
        title_col = idx.get("活動名稱 / Activity Name")
        loc_col = idx.get("地點 / Location")
        if not title_col or not loc_col:
            continue

        current_event_row = None
        for row in range(2, ws.max_row + 1):
            title = str(ws.cell(row, title_col).value or "").strip()
            loc_cell = ws.cell(row, loc_col)
            loc_value = loc_cell.value

            if title and loc_value:
                current_event_row = row
                continue

            if not title and loc_value and current_event_row and looks_like_addr(loc_value):
                prev = ws.cell(current_event_row, loc_col)
                merged = merge_location(prev.value, loc_value)
                if merged != norm(str(prev.value or "")):
                    prev.value = merged
                    total_merged += 1
                    details.append({"sheet": ws.title, "event_row": current_event_row, "address_row": row, "location": merged})
                if not row_has_meaningful_data(ws, row, loc_col):
                    loc_cell.value = None
                    total_cleared += 1

    if total_merged:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = os.path.join(BACKUP_DIR, f"全台ACG活動.before_clean_{stamp}.xlsx")
        shutil.copy2(XLSX, backup)
        wb.save(XLSX)
    print(json.dumps({
        "merged": total_merged,
        "cleared_address_rows": total_cleared,
        "backup_created": bool(total_merged),
        "examples": details[:8],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
